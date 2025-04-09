from flask import Flask, request, send_file, jsonify, render_template
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
from io import BytesIO
import openpyxl
import os

app = Flask(__name__)

# 密碼設定
USER_PASSWORD = os.environ.get("USER_PASSWORD", "user123")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")

# PostgreSQL 連線
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get("DATABASE_URL")
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# 資料表 schema
class Revision(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    category = db.Column(db.String(100), nullable=False)
    before = db.Column(db.Text, nullable=False)
    after = db.Column(db.Text, nullable=False)
    timestamp = db.Column(db.String(20), nullable=False)

with app.app_context():
    db.create_all()

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/admin")
def admin():
    return render_template("admin.html")

@app.route("/submit", methods=["POST"])
def submit():
    password = request.form.get("password")
    category = request.form.get("category")
    before = request.form.get("before")
    after = request.form.get("after")

    if password != USER_PASSWORD:
        return jsonify({"error": "使用者密碼錯誤"}), 401
    if not category or not before or not after:
        return jsonify({"error": "請完整填寫類別、修改前、修改後"}), 400

    exists = Revision.query.filter_by(category=category, before=before, after=after).first()
    if exists:
        return jsonify({"error": "已存在相同資料"}), 409

    revision = Revision(
        category=category,
        before=before,
        after=after,
        timestamp=datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    )
    db.session.add(revision)
    db.session.commit()
    return jsonify({"message": "提交成功"})

@app.route("/upload_excel", methods=["POST"])
def upload_excel():
    password = request.form.get("password")
    file = request.files.get("excel_file")

    if password != USER_PASSWORD:
        return jsonify({"error": "使用者密碼錯誤"}), 401
    if not file:
        return jsonify({"error": "請選擇檔案"}), 400

    wb = openpyxl.load_workbook(file)
    ws = wb.active

    count = 0
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        category, before, after = row
        if not category or not before or not after:
            continue
        exists = Revision.query.filter_by(category=category, before=before, after=after).first()
        if not exists:
            revision = Revision(
                category=category,
                before=before,
                after=after,
                timestamp=datetime.now().strftime("%Y/%m/%d %H:%M:%S")
            )
            db.session.add(revision)
            count += 1

    db.session.commit()
    return jsonify({"message": f"成功合併 {count} 筆資料"})

@app.route("/download", methods=["POST"])
def download_excel():
    password = request.form.get("password")
    if password not in [USER_PASSWORD, ADMIN_PASSWORD]:
        return jsonify({"error": "密碼錯誤"}), 401

    revisions = Revision.query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "修訂記錄"

    # 欄位標題與欄寬設定
    headers = ["類別", "修改前", "修改後", "時間戳記"]
    ws.append(headers)
    column_widths = [12, 24, 36, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # 凍結首列
    ws.freeze_panes = "A2"

    # 加入資料與邊框
    border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )

    for r in revisions:
        ws.append([r.category, r.before, r.after, r.timestamp])

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=4):
        for cell in row:
            cell.border = border

    # 回傳 Excel
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="output.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/clear", methods=["POST"])
def clear_excel():
    password = request.form.get("password")
    if password != ADMIN_PASSWORD:
        return jsonify({"error": "管理者密碼錯誤"}), 401

    num_deleted = db.session.query(Revision).delete()
    db.session.commit()
    return jsonify({"message": f"資料已清空，共刪除 {num_deleted} 筆資料"})

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
